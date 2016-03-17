#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl as pyxl
""" 
Add results of fit to .XLS worksheet

"""

def write_xls(xls_filename, fit_name, E0, W):
    
    wb = pyxl.load_workbook(xls_filename)
    ws = wb.active
#==============================================================================
#     max_row = ws.max_row
#     
#     
#     E0 = float(0.93)
#     W = 0.3
#     ws.cell(row=max_row+1, column = 1).value = E0
#     print(ws.cell(row=max_row+1, column = 1).value)
#     ws.append(['fit001',E0,W])
#==============================================================================
    
    i_found = None
    row_to_write = None
    
    for i in range(1, ws.max_row+1):
        if ws.cell(row=i,column=1).value == fit_name:
            i_found = i
            
    if i_found:
        print('An entry for fit', fit_name, 'already exists')
        
        while True:        
            ans = input('Overwrite (O)  Append (A)  or Skip (S): ').upper()
            if ans.startswith('O'):
                row_to_write = i_found
                break
            elif ans.startswith('A'):
                row_to_write = ws.max_row+1
                break
            elif ans.startswith('S'):
                row_to_write=None
                break
            else:
                print('Please enter "O", "A", or "S" : ')
                
    if(row_to_write):
        ws.cell(row=row_to_write, column = 1).value = fit_name
        ws.cell(row=row_to_write, column = 2).value = E0
        ws.cell(row=row_to_write, column = 3).value = W
    
    
    wb.save(xls_filename)


if __name__ == '__main__':

    write_xls('fits\\fit_results.xlsx', 'fit001', 2, 2*3.1415927)