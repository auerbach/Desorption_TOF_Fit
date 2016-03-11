#!/usr/bin/env python
""" 
Desorption_TOF_Fit
Fit data from post permeation desorption TOF
"""
#==============================================================================
# 
# Based on Time of Flight fitting program
#   Alessandro Genova, Leiden University, July 2012 
#   Francesco Nattino, Leiden University, June 2015  (vs 7)
#
#==============================================================================

from lmfit import minimize, fit_report
import numpy as np
import openpyxl as pyxl
from scipy import special
import shutil
import subprocess
import unicodedata
# from glob import glob

from parse_command_file import parse_cmd_file
from PlotFit import PlotFit
from Cutoff import cutoff_function
from compute_tof import TOF
import TOF_fit_global
from Parameters2 import Parameters2
# from Parameters2 import Parameter2, Parameters2

#import global_variables_old as glbl_old
from Data import Data

def GenerateThetaAngles(AveragingType, GridType,        \
                        NPointsSource, NPointsDetector, \
                        ZSource,    RSource,            \
                        ZAperture, RAperture,           \
                        ZDetector,  LengthDetector):

    if AveragingType == 'PointDetector':
        ThetaAngles = np.arange( 0., glbl.AngRes + glbl.ThetaStep, glbl.ThetaStep )
    elif AveragingType == 'None':
        ThetaAngles = [0.]
    elif AveragingType == 'LineDetector':
        import GeneratePoints
        GridOfPointsSource  = GeneratePoints.PointsOnTheSource(         \
            GridType = GridType, ZSource = ZSource , RSource = RSource, \
            NPoints = NPointsSource)

        GridOfPointsDetector = GeneratePoints.PointsOnTheDetectionLine( \
            ZDetector = ZDetector ,         \
            NPoints = NPointsDetector,      \
            Length = LengthDetector )

        ThetaAngles = GeneratePoints.ThetaPossibleTrajectories(         \
            GridSource   = GridOfPointsSource,                          \
            GridDetector = GridOfPointsDetector,                        \
            ZAperture = ZAperture,                                      \
            RAperture = RAperture)


        for i in range( len( ThetaAngles )):
            ThetaAngles[i] = np.degrees( ThetaAngles[i] )
        print("Considering ", len(ThetaAngles ),\
            " values of Theta in the angular averaging, minimum: %8.3f"\
            %min( ThetaAngles), " deg , maximum: %8.3f" %max( ThetaAngles) ," deg.")
    return ThetaAngles




# -------------------------------------------------------------------------------------------------
#   FitData -- function to perform the fit
# -------------------------------------------------------------------------------------------------
def FitData( DataSets, Params, AveragingType, ProbCurveType, mass_molecules):
    # Fit the data
    # Give to the datasets a form that "minimize" likes
    X, Y = [], []

      
    for DataSet in DataSets:
        X = X + list( DataSet[0] )
        Y = Y + list( DataSet[1] )
        
    # Perform NLLSQ fit
    result = minimize( Residual, Params, 
                      args=(X, Y, DataSets, AveragingType, ThetaAngles, 
                            ProbCurveType, mass_molecules) )
    
    return result

    
# -------------------------------------------------------------------------------------------------
#   Residual -- residual function used in fit
# -------------------------------------------------------------------------------------------------
def Residual( Params, X, Y, DataSets, AveragingType, ThetaAngles, ProbCurveType, mass_molecules):
    # Residual function needed by minimize
    Resid = []

    for i in range( len( DataSets )):
        NDataSet = i + 1
        Resid = Resid + list(DataSets[i][1] \
                - TOF(DataSets[i][0], NDataSet, Params, data, glbl, AveragingType, ThetaAngles,
                      ProbCurveType, cutoff_type, mass_molecules))

    return np.array( Resid )
    


def ProbFromTOFInversion(Time, Signal, NDataSet, Params, AveragingType, ThetaAngles, 
                         ProbCurveType, cutoff_type, mass_molecules):
    # Time of flight signal model. The function takes an uncorrected time in seconds and returns a signal
    
    mass_factor = np.sqrt(mass_molecules[NDataSet -1] / glbl.massH2)
    FFRDist  = Params['FFR_%i'      %NDataSet].value * 1E-3
    MaxTOF   = Params['Yscale_%i'   %NDataSet].value
    Baseline = Params['Baseline_%i' %NDataSet].value
    TimeCorr = Params['IonTOF_%i'   %NDataSet].value * mass_factor
    Temperature = Params['Temp_%i'  %NDataSet].value

    # subtract the ion flight time and eliminate singularity that would occur at Time = 0    
    Time = Time - TimeCorr * 1E-6  
    Time = np.where(Time != 0, Time, np.repeat(0.01E-6, len(Time)))
            
    CutOff = cutoff_function(Params, data, NDataSet, Time, cutoff_type)
    
#==============================================================================
#     if cutoff_type.lower == 'tanh':
#         TCutC    = Params['TCutC_%i'     %NDataSet].value
#         TCutW    = Params['TCutW_%i'     %NDataSet].value
#         CutOff = 0.5 * (1. - np.tanh((Time - TCutC*1E-6) / (TCutW*1E-6)))
#     elif cutoff_type.lower == 'exp':
#         FFRDist  = Params['FFR_%i'     %NDataSet].value * 1E-3
#         ECutM    = Params['ECutM_%i' %NDataSet].value
#         ECutS    = Params['ECutS_%i' %NDataSet].value
#         mass = data.mass_molecules[NDataSet-1]
#         Velocity = FFRDist / Time 
#         Ekin = (0.5 * mass * Velocity**2.) * glbl.eVConst
#         if Ekin > ECutM:
#             CutOff = 1.0 - np.exp(-ECutS * (Ekin - ECutM))
#         else:
#             CutOff = 0.0
#==============================================================================

    VelocityDistribution = 0.                        # Initialize Signal to 0

    if AveragingType == "PointDetector":    
        # Averaging performed taking into account different flight time for different angles
        for Theta in ThetaAngles:
            # v = x / t = ( L / cos(theta) ) / t
            Velocity = FFRDist /(Time * np.cos(np.radians(Theta))) 
            Ekin = (0.5 * glbl.massAmu * Velocity ** 2.) * glbl.eVConst
            # Enorm = Ekin * np.cos( np.radians(Theta) )**2. # Reaction probability depends on normal energy
            VelocityDistribution = VelocityDistribution                   \
                                   + Velocity**4.                         \
                                   * np.exp(-Ekin / (glbl.kb * Temperature)) \
                                   * np.cos( np.radians(Theta) )**2.      \
                                   * np.sin( np.radians(Theta) ) \
                                     * glbl.ThetaStep

    elif AveragingType == "None":
        # No angular averaging
        Velocity = FFRDist / Time
        Ekin = (0.5 * glbl.massAmu * Velocity ** 2.) * glbl.eVConst
        # Enorm = Ekin
        VelocityDistribution = Velocity**4. * np.exp(-Ekin / (glbl.kb * Temperature))
        
        # kludge to fix divide by zero runtime error        
        for ii in range(len(VelocityDistribution)):
            if VelocityDistribution[ii] == 0:
                VelocityDistribution[ii] = 1E-100
        

    ProbFromTOF = (Signal - Baseline ) / (MaxTOF * VelocityDistribution * CutOff)
    Energy = 0.5 * glbl.massAmu * (FFRDist / Time) ** 2. * glbl.eVConst
    

    return Energy, ProbFromTOF


#==============================================================================
#     TOFInvertedOutFile         = "DataSet_" + str( NDataSet ) + "_" + State + "_" + Label + "_ReacProb_Fit_" + ProbCurveType + "_FromTOFInversion.dat"
#     # Reaction probability curve: minimum energy, maximum energy, deltaE (eV)
#     Emin = 0.0
#     Emax = 2.0
#     DeltaE = 0.01
# 
#     Energy = np.arange( Emin, Emax + DeltaE, DeltaE)
# 
#     # Calculate Reaction probability fitted curve
#     ReactionProbability = Prob(Energy, NDataSet, Params, ProbCurveType)
#     np.savetxt( ReactionProbabilityOutFile, np.column_stack(( Energy, ReactionProbability)))
#     # print("WriteProbOutput: Reaction probability curve written to file ", ReactionProbabilityOutFile)
# 
#     # Reaction probability from TOF inversion; only possible with Point detector or no angular averaging (and normal energy scaling!)
#     if AveragingType != "LineDetector":
#         TOFEnergy, TOFReactionProbability = ProbFromTOFInversion(TOFTime, TOFData, NDataSet, Params, AveragingType, ThetaAngles, ProbCurveType, cutoff_type)
#         np.savetxt( TOFInvertedOutFile, np.column_stack(( TOFEnergy, TOFReactionProbability*Params['Yscale_%i' %NDataSet].value, Prob( TOFEnergy, NDataSet, Params, ProbCurveType)*Params['Yscale_%i' %NDataSet].value )))
#         # print("WriteProbOutput: Inverted TOF written to file ", TOFInvertedOutFile)
#         # print('WriteProbOutput: Exiting')
#    return                                               
#==============================================================================









#============================================================================= 
#============================================================================= 
# ============================  Main PROGRAM  ================================
#============================================================================= 
#============================================================================= 

import time
start_time = time.time()

#------------------------------------------------------------------------------
# Create TOF_fit_global, Data, and Parameters, objects
#------------------------------------------------------------------------------
glbl = TOF_fit_global.TOF_fit_global()
data = Data(glbl)
parms = Parameters2()

# define default path to control files and fit output
pathToFits = 'Fits\\'
# pathToFits = 'd:\\users\\dja\\desktop\\permeation\\All-DJA\\Fits'

#------------------------------------------------------------------------------
# uncomment for testing:
#------------------------------------------------------------------------------
newFitNumber = '034'
newFitNumber = '017'

#==============================================================================
# #------------------------------------------------------------------------------
# # begin1 comment out for testing
# #------------------------------------------------------------------------------
# # Get Last fit number from FitNumber.dat
# fit_number_file = open(pathToFits + 'FitNumber.dat', 'r+')
# fitNumber = '{:03d}'.format(int(fit_number_file.readline()))
# oldFitNumber = fitNumber
# newFitNumber = fitNumber
# 
# while True:
#     print('please enter oldfit number: ', '[', oldFitNumber, ']')
#     ans = input('?')
#     if ans:
#         old_n = '{:03d}'.format(int(ans))
#     else:
#         old_n = oldFitNumber
#     
#     if int(old_n) > int(oldFitNumber):
#         print('maximum allowed for old fit number is ', oldFitNumber)
#     else:
#         break
#     
# oldFitNumber = old_n
# 
# ans = input ('make new command file? [no]')
# if ans:
#     if ans.upper()[0] == 'Y':
#         newFitNumber = '{:03d}'.format(int(fitNumber)+1)
#     fit_number_file.seek(0)
#     fit_number_file.write(newFitNumber)
# else:
#     newFitNumber = oldFitNumber
# 
# fit_number_file.close()
# 
# oldFile = pathToFits + 'Fit' + oldFitNumber + '.tof_in'
# newFile = oldFile
# 
# if oldFitNumber != newFitNumber:
#     newFile = pathToFits + 'Fit' + newFitNumber + '.tof_in'
#     shutil.copy2(oldFile, newFile)
# 
# subprocess.call(['npp.bat', newFile])
# 
# #------------------------------------------------------------------------------
# # end of comment out for testing
# #------------------------------------------------------------------------------
#==============================================================================

cmdFile = pathToFits + 'Fit' + str(newFitNumber) + '.tof_in'


#==============================================================================
# # Parse the command file
#==============================================================================
errors = parse_cmd_file(cmdFile, glbl, parms)


if len(errors) > 0:
    print('Errors\n', errors)
    raise SystemExit ('Error in command file' )

#==============================================================================
# # use Francesco's names until have time to change
#==============================================================================
DataFiles = glbl.signalFiles
BackgroundFiles = glbl.backgroundFiles

# initalize lists
states = []         # list if tuples (v, j)
temperatures = []
DataSets = []
PlotDataSets = []

# Read data from input
fit_ranges = []
for i in range( len( DataFiles)):
    Tmin = glbl.Tmins[i]
    Tmax = glbl.Tmaxs[i]
    ProbCurveType = glbl.Functions[i]
    DataFile = DataFiles[i]
    AveragingType = glbl.AveragingTypes[i]
    cutoff_type = glbl.cutoff_types[i]
    
    if BackgroundFiles  != [ "" ]:
        BackgroundFile = BackgroundFiles[i]
    else:
        BackgroundFile = BackgroundFiles[0]
          
    data.read_data(DataFile, BackgroundFile, Tmin, Tmax)  
    
    Nmin = data.fit_ranges[i][0]
    Nmax = data.fit_ranges[i][1]    
    Tmin = data.datasets[i][0][Nmin]
    Tmax = data.datasets[i][0][Nmax]
    
    fit_ranges.append([Nmin,Nmax])    
    states.append( data.states[i] )
    temperatures.append(data.temperatures[i])
    
    Time   = data.datasets[i][0]
    Signal = data.datasets[i][1]
    DataSets.append([Time[Nmin:Nmax], Signal[Nmin:Nmax]])
    PlotDataSets.append([Time, Signal])
    
# Generate Theta angles employed for angular averaging
ThetaAngles = GenerateThetaAngles(
                AveragingType=AveragingType,
                GridType=glbl.GridType,
                NPointsSource=glbl.NPointsSource,
                NPointsDetector=glbl.NPointsDetector,
                ZSource = glbl.ZSource,
                RSource = glbl.RSource,
                ZAperture = glbl.ZAperture, RAperture = glbl.RAperture,
                ZDetector = glbl.ZLaser, LengthDetector = glbl.LLaser)
                # ZDetector = ZFinal, LengthDetector = 2.*RFinal         \


#--------------------------------------------------------------------------------------------------
# Fit the data to model 
#--------------------------------------------------------------------------------------------------
fitResult = FitData( DataSets, parms, AveragingType,
                    ProbCurveType, data.mass_molecules)
print(fit_report(fitResult))

state_string = 'v' + str(states[0][0]) + 'j' +str(states[0][1])
result_file_name = 'Fit' + newFitNumber + '_' + data.molecules[0] + \
                   '_' + state_string + '_' + ProbCurveType + '.fit_out'

# begin comment out for profiling
with open(pathToFits + result_file_name, 'w') as result_file:
    result_file.write('#\n')
    result_file.write('# Fit {} Results\n'.format(newFitNumber))
    result_file.write('#\n')
    result_file.write('#' + 60*'-' + '\n')
    result_file.write('# Control file: {}\n'.format(cmdFile))
    result_file.write('#' + 60*'-' + '\n')
    
    #----------------------------------------------------------------------------------------------
    # write the control file to the result file
    #----------------------------------------------------------------------------------------------
    with open(cmdFile,'r') as cmd:
        cmd_lines = cmd.readlines()

    for line in cmd_lines:
        result_file.write('# ' + line )
    
    result_file.write('\n#' + 60*'-' + '\n')    
    result_file.write('# End Control File\n')
    result_file.write('#' + 60*'-' + '\n')
    result_file.write('\n')
    
    #----------------------------------------------------------------------------------------------
    # write the angular averaging parameters
    #----------------------------------------------------------------------------------------------
    result_file.write('#' + 60*'-' + '\n')
    result_file.write('# Angular Averaging Parameters\n')
    result_file.write('#' + 60*'-' + '\n')
    result_file.write('# NPointsSource   : ' + str(glbl.NPointsSource) + '\n')
    result_file.write('# NPointsDetector : ' + str(glbl.NPointsDetector) + '\n')
    result_file.write('# ZSource         : ' + str(glbl.ZSource) + '\n')
    result_file.write('# RSource         : ' + str(glbl.RSource) + '\n')
    result_file.write('# ZAperture       : ' + str(glbl.ZAperture) + '\n')
    result_file.write('# RAperture       : ' + str(glbl.RAperture) + '\n')
    result_file.write('# ZFinal          : ' + str(glbl.ZFinal) + '\n')
    result_file.write('# RFinal          : ' + str(glbl.RFinal) + '\n')
    # result_file.write('# it mse\n')
    
    for i in range(0, len(ThetaAngles), 10):
        result_file.write('# ThetaAngles : ' + str(ThetaAngles[i : i+10]) + '\n')
    result_file.write('#' + 60*'-' + '\n')    
    result_file.write('# End Angular Averaging Parameter\n')
    result_file.write('#' + 60*'-' + '\n')
    result_file.write('\n')
    
    #----------------------------------------------------------------------------------------------
    # write the fit report to the result file
    #----------------------------------------------------------------------------------------------
    result_file.write('#' + 60*'-' + '\n')
    result_file.write('# Fit Report\n')
    result_file.write('#' + 60*'-' + '\n')    

    report = fit_report(fitResult).split('\n')
    for line in report:
        result_file.write('# ' + line + '\n')
    result_file.write('#' + 60*'-' + '\n')    
    result_file.write('# End Fit Report\n')
    result_file.write('#' + 60*'-' + '\n')
    result_file.write('\n')
    
    #----------------------------------------------------------------------------------------------
    # the number of datasets to plot
    #----------------------------------------------------------------------------------------------
    result_file.write('#' + 60*'-' + '\n')
    result_file.write('# Labels and data for plots\n')
    result_file.write('#' + 60*'-' + '\n')
    result_file.write('# Number of plots : ' + str(len(DataSets)) + '\n')
    result_file.write('#' + 60*'-' + '\n')
    result_file.write('\n')
    
    #----------------------------------------------------------------------------------------------
    # for each data set, write plot title, labels, and data
    #----------------------------------------------------------------------------------------------
    for i in range( len( DataSets )):
        n_dataset = i + 1
        
        result_file.write('#' + 60*'-' + '\n')
        result_file.write('# Plot ' + str(n_dataset) + '\n')
        result_file.write('#' + 60*'-' + '\n')
        
        #------------------------------------------------------------------------------------------
        # write the plot title
        #------------------------------------------------------------------------------------------
        result_file.write('# Title    : Fit {:03d}_{}'.format(int(newFitNumber), n_dataset) + '\n')

        #------------------------------------------------------------------------------------------
        # write the plot label
        #------------------------------------------------------------------------------------------
        result_file.write('# Label    : ' + data.original_signal_names[i] + '\n')
        result_file.write('# Label    : \n')
        # result_file.write('# Label    : ' + '----------------------\n')
        
        if AveragingType == 'None' : 
            avg_type_label = 'No Angular Averaging'
        elif AveragingType == 'PointDetector':
            avg_type_label = 'Point Detector'
        elif AveragingType == 'LineDetector':
            avg_type_label = 'Line Detector' 
        
        # result_file.write('# Label    : Averaging: ' + avg_type + '\n')
        result_file.write('# Label    : Function:  ' + ProbCurveType + '\n')
        result_file.write('# Label    : ' + avg_type_label + '\n')
        result_file.write('# Label    : \n')         
        final_params = fitResult.params     
        
        parm_for_plot_label = ['E0', 'W', 'TCutC', 'TCutW', 'ECutM', 'ECutS', 'FFR']
        # if glbl.Functions[i].lower().startswith('cal'):
        #     parm_for_plot_label.append('FFR')
        for p in parm_for_plot_label:
            p_ = p + '_' + str(i+1)
            try:
                result_file.write('# Label    : {:6s}{:>7.3f}'.format(p, final_params[p_].value))
                
                if parms[p +'_1'].glbl and i >0:
                    result_file.write(' (global)\n')
                if not final_params[p_].vary:
                    result_file.write(' (fixed)\n')
                else:
                    stderr = '{:6.3f}'.format(final_params[p_].stderr)
                    result_file.write(' (' + unicodedata.lookup('plus-minus sign') 
                                      + str(stderr) +')\n')
            except:
                pass
        
        #------------------------------------------------------------------------------------------
        # Get the point where time > 2E-6 sec and ignore earlier points to avoid noise
        #------------------------------------------------------------------------------------------
        i_start = np.where(PlotDataSets[i][0] > 2.0E-6)[0][0]        
        
        #------------------------------------------------------------------------------------------
        # Get the plot data and convert time to microseconds
        #-----------------------------------------------------------------------------------------
        Time     =  PlotDataSets[i][0][i_start:] 
        Signal   =  PlotDataSets[i][1][i_start:]
        Fit      =  TOF(Time, n_dataset, fitResult.params, data, glbl, AveragingType, ThetaAngles,
                        ProbCurveType, cutoff_type, data.mass_molecules)
        
                
        mass_factor = np.sqrt(data.mass_molecules[i] / glbl.massH2)
        ion_tof  = fitResult.params['IonTOF_%i'   %n_dataset].value * mass_factor 
        Time2 = Time - ion_tof * 1E-6  
        Time2 = np.where(Time2 != 0, Time2, np.repeat(0.01E-6, len(Time)))
        Cutoff   =  cutoff_function(fitResult.params, data, glbl, n_dataset, Time2, cutoff_type )
                                    
        Time = Time * 1E6    # convert to microseconds for plotting
        
        #------------------------------------------------------------------------------------------
        # write the number of data lines and range of lines fit
        #------------------------------------------------------------------------------------------
        result_file.write('# Npoints    : ' + str(len(Time)) +'\n')
        result_file.write('# Nmin, Nmax : ' + str(Nmin) + ',' +str(Nmax) + '\n')
        result_file.write('# Tmin, Tmax : ' + str(Tmin * 1E6) + ',' +str(Tmax * 1E6) + '\n')
        result_file.write('# Baseline   : ' + 
                            str(fitResult.params['Baseline_' + str(n_dataset)].value) +'\n')
               
        result_file.write('#' + 68*'-' + '\n')
        
        #------------------------------------------------------------------------------------------
        # write the plot data
        #------------------------------------------------------------------------------------------
        result_file.write('# Time            Signal                Fit                 Cutoff\n')
        result_file.write('#' + 68*'-' + '\n')
        result_file.write('# Begin data\n')
        
        
        for j in range(len(Time)):
            result_file.write('{:6.2f} {:20.5e} {:20.5e} {:20.5e}\n'
                                .format(Time[j], Signal[j], Fit[j], Cutoff[j]))
        
        result_file.write('# End data\n')
        result_file.write('#' + 68*'-' + '\n')
        result_file.write('# End Plot ' + str(n_dataset) +'\n')
            
# end comment out for profiling        

print("--- %s seconds ---" % (time.time() - start_time))

PlotFit(pathToFits + result_file_name)
#==============================================================================
# # Begin comment out for testing
#
# #--------------------------------------------------------------------------------------------------
# # Enter fit results in results.xlsx
# #--------------------------------------------------------------------------------------------------
# xls_filename ='Fits\\fit_results.xlsx'
# wb = pyxl.load_workbook(xls_filename)
# ws = wb.active 
#    
# i_found = None
# next_row = None
# fit_name = 'Fit' + newFitNumber
# 
# for i in range(1, ws.max_row+1):
#     test = ws.cell(row=i, column=1).value
#     try:
#         if ws.cell(row=i,column=1).value.startswith(fit_name):
#             i_found = i
#     except:
#         pass
#         
# if i_found:
#     print('An entry for fit', fit_name, 'already exists')
#     
#     while True:        
#         ans = input('Overwrite (O)  Append (A)  or Skip (S): ').upper()
#         if ans.startswith('O'):
#             next_row = i_found
#             break
#         elif ans.startswith('A'):
#             next_row = ws.max_row+1
#             break
#         elif ans.startswith('S'):
#             next_row=None
#             break
#         else:
#             print('Please enter "O", "A", or "S" : ')
#             
# else: 
#     next_row = ws.max_row + 2
#             
# if(next_row):
#     for n in range(len(glbl.signalFiles)):
#         ws.cell(row=next_row, column = 1).value = fit_name + '_' + str(n+1)
#         ws.cell(row=next_row, column = 2).value = data.molecules[n]  # molecule
#         ws.cell(row=next_row, column = 3).value = data.states[n][0]
#         ws.cell(row=next_row, column = 4).value = data.states[n][1]
#         ws.cell(row=next_row, column = 5).value = glbl.Functions[n]
#         ws.cell(row=next_row, column = 6).value = glbl.AveragingTypes[n]
#         if glbl.Functions[0].lower() == 'erf':
#             ws.cell(row=next_row, column = 7).value = fitResult.params['E0_' + str(n+1)].value
#             ws.cell(row=next_row, column = 8).value = fitResult.params['E0_' + str(n+1)].stderr
#             ws.cell(row=next_row, column = 9).value = fitResult.params['W_'  + str(n+1)].value
#             ws.cell(row=next_row, column =10).value = fitResult.params['W_'  + str(n+1)].stderr
#         ws.cell(row=next_row, column = 11).value = glbl.Tmins[0]
#         ws.cell(row=next_row, column = 12).value = glbl.Tmaxs[0]
#         ws.cell(row=next_row, column = 13).value = fitResult.params['FFR_'   + str(n+1)].value
#         ws.cell(row=next_row, column = 14).value = fitResult.params['ECutM_' + str(n+1)].value 
#         ws.cell(row=next_row, column = 15).value = fitResult.params['ECutS_' + str(n+1)].value
#         if n == 0:
#             ws.cell(row=next_row, column = 16).value = glbl.comment_xlsx
#         next_row += 1
# wb.save(xls_filename)
#
# end comment out for testing
#==============================================================================




    # if not calibration run
#==============================================================================
#     if not glbl.Functions[i].lower().startswith('cal'):
#         WriteProbOutput(Time, Signal, State, Label, NDataSet, fitResult.params, 
#                         AveragingType, ThetaAngles, ProbCurveType)                                                   
#==============================================================================
