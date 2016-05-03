#!/usr/bin/env python
"""
functions to write .fit_out, .plot_out, and summary.xls

write_fit_out(glbl, data, path, control_filename, 
              fit_number, fit_result, PlotDataSets)
    glbl        glbl variables and constants class
    data        data class
    path        path to the control file and output files
    control_filename name of the .fit_in control file
    fit_result  minimizer object with the fit results
    plot_data_sets  data to use for output
        three collumns time, signal, 
    

"""

import lmfit
import numpy as np
import unicodedata
import openpyxl as pyxl
import compute_tof
from cutoff import cutoff_function


def write_fit_out(glbl, data, path, control_filename, fit_number, fit_result, PlotDataSets):
    state_string = 'v' + str(data.states[0][0]) + 'j' + str(data.states[0][1])
    result_filename = 'Fit' + fit_number + '_' + data.molecules[0] + \
                      '_' + state_string + '_' + glbl.functions[0] + '.fit_out'

    parms = fit_result.params

    with open(path + result_filename, 'w') as result_file:
        result_file.write('#\n')
        result_file.write('# Fit {} Results\n'.format(fit_number))
        result_file.write('#\n')
        result_file.write('#' + 60 * '-' + '\n')
        result_file.write('# Control file: \n') 
        result_file.write('# {}\n'.format(control_filename))
        result_file.write('#' + 60 * '-' + '\n')

        # --------------------------------------------------------------------------
        # write the control file to the result file
        # --------------------------------------------------------------------------
        with open(control_filename, 'r') as cmd:
            cmd_lines = cmd.readlines()

        for line in cmd_lines:
            result_file.write('# ' + line)

        # result_file.write('\n#' + 60 * '-' + '\n')
        # result_file.write('# End Control File\n')
        # result_file.write('#' + 60 * '-' + '\n')

        # --------------------------------------------------------------------------
        # write the angular averaging parameters
        # --------------------------------------------------------------------------
        result_file.write('#\n')
        result_file.write('#' + 60 * '-' + '\n')
        result_file.write('# Angular Averaging Parameters\n')
        result_file.write('#' + 60 * '-' + '\n')
        result_file.write('# points_source    : ' + str(glbl.points_source) + '\n')
        result_file.write('# points_detector  : ' + str(glbl.points_detector) + '\n')
        result_file.write('# z_source         : ' + str(glbl.z_source) + '\n')
        result_file.write('# r_source         : ' + str(glbl.r_source) + '\n')
        result_file.write('# z_aperture       : ' + str(glbl.z_aperture) + '\n')
        result_file.write('# r_aperture       : ' + str(glbl.r_aperture) + '\n')
        result_file.write('# z_final          : ' + str(glbl.z_final) + '\n')
        result_file.write('# r_final          : ' + str(glbl.r_final) + '\n')
        # result_file.write('# it mse\n')

        for i in range(0, len(glbl.angles_list), 10):
            result_file.write('# angles_list : ' + str(glbl.angles_list[i: i + 10]) + '\n')
        # result_file.write('#' + 60 * '-' + '\n')
        # result_file.write('# End Angular Averaging Parameter\n')
        # result_file.write('#' + 60 * '-' + '\n')
        result_file.write('#\n')

        # --------------------------------------------------------------------------
        # write the fit report to the result file
        # --------------------------------------------------------------------------
        result_file.write('#' + 60 * '-' + '\n')
        result_file.write('# Fit Report\n')
        result_file.write('#' + 60 * '-' + '\n')

        #------------------------------------------------------------------------------------------
        # first write chi square and reduced chi square in scientific format
        # standard fit report use numbers with 3 places after decimal point
        # which results in print 0.000 sometimes
        #------------------------------------------------------------------------------------------
        result_file.write('# chi-square             = {:10.3e}'.format(fit_result.chisqr) + '\n')
        result_file.write('# reduced chi-square     = {:10.3e}'.format(fit_result.redchi) + '\n')
        result_file.write('#\n')
        
        report = lmfit.fit_report(fit_result).split('\n')
        for line in report:
            result_file.write('# ' + line + '\n')
        # result_file.write('#' + 60 * '-' + '\n')
        # result_file.write('# End Fit Report\n')
        # result_file.write('#' + 60 * '-' + '\n')
        result_file.write('#\n')

        # --------------------------------------------------------------------------
        # the number of datasets to plot
        # --------------------------------------------------------------------------
        result_file.write('#' + 60 * '-' + '\n')
        result_file.write('# Data for plots\n')
        result_file.write('#' + 60 * '-' + '\n')
        result_file.write('# Number of plots : ' + str(len(PlotDataSets)) + '\n')
        result_file.write('#' + 60 * '-' + '\n')
        result_file.write('#\n')

        # --------------------------------------------------------------------------
        # for each data set, write plot title, information for labels, and data
        # --------------------------------------------------------------------------
            
        for i in range(len(PlotDataSets)):
            n_dataset = i + 1

            n_min = data.fit_index_ranges[i][0]
            n_max = data.fit_index_ranges[i][1]
            t_min = data.datasets[i][0][n_min] * 1E6
            t_max = data.datasets[i][0][n_max] * 1E6
            ProbCurveType = glbl.functions[i]
            averaging_type = glbl.averaging_types[i]
            # DataFile = signal_filenames[i]
            cutoff_type = glbl.cutoff_types[i]

            result_file.write('#' + 60 * '-' + '\n')
            result_file.write('# Plot ' + str(n_dataset) + '\n')
            result_file.write('#' + 60 * '-' + '\n')

            # --------------------------------------------------------------------------
            # write the plot title
            # --------------------------------------------------------------------------
            result_file.write('# Title             : ')
            result_file.write('Fit {:04d}_{}'.format(int(fit_number), n_dataset) + '\n')

            # -------------------------------------------------------------------------------------
            # write information for the plot labels
            # -------------------------------------------------------------------------------------
            comment = glbl.comments[i]
            if comment == None:
                comment = ''
            
            signal_file = glbl.signal_filenames[i]
            
            background_file = glbl.background_filenames[i]
            if background_file == None:
                background_file = 'none'
            
            result_file.write('# commment          : ' + comment             + '\n')
            result_file.write('# signal_file       : ' + signal_file         + '\n')
            result_file.write('# background_file   : ' + background_file     + '\n')
            result_file.write('# date              : ' + data.dates[i]       + '\n')            
            result_file.write('# molecule          : ' + data.molecules[i]   + '\n')
            result_file.write('# states            : ' + str(data.states[i]) + '\n')
            
            result_file.write('# function          : ' + ProbCurveType       + '\n')
            
            if averaging_type == 'none':
                avg_type_label = 'No Averaging'
            elif averaging_type == 'point_detector':
                avg_type_label = 'Point Detector'
            elif averaging_type == 'line_detector':
                avg_type_label = 'Line Detector'
            result_file.write('# avg_type_label    : ' + avg_type_label + '\n')
            
            #--------------------------------------------------------------------------------------
            # write time range of fit and t_min, t_max, threshold, number points averaged
            #--------------------------------------------------------------------------------------
            result_file.write('#\n')
            result_file.write('#' + 60 * '-' + '\n')
            result_file.write('# fit range and how determined \n')
            result_file.write('#' + 60 * '-' + '\n')
            n_tup = (n_min, n_max)
            t_tup = (t_min, t_max)
            result_file.write('# fit_time_range    : ' + str(glbl.fit_ranges[i])      + '\n')            
            result_file.write('# n_min, n_max      : ' + str(n_tup)                   + '\n')
            result_file.write('# threshold         : ' + str(glbl.thresholds[i])      + '\n')
            result_file.write('# n_delt for avg    : ' + str(glbl.n_delts[i])          + '\n')
            result_file.write('# t_min, t_max      : ' + str(t_tup)                   + '\n')
            
            #--------------------------------------------------------------------------------------
            # write time range of fit and t_min, t_max, threshold, number points averaged
            #--------------------------------------------------------------------------------------
            result_file.write('#\n')
            result_file.write('#' + 60 * '-' + '\n')
            result_file.write('# baseline and how determined \n')
            result_file.write('#' + 60 * '-' + '\n')
            baseline = fit_result.params['baseline_' + str(n_dataset)].value            
            result_file.write('# baseline_range    : ' + str(glbl.baseline_ranges[i]) + '\n')            
            result_file.write('# baseline          : ' + str(baseline)                + '\n')
            result_file.write('#\n')
            
            #--------------------------------------------------------------------------------------
            # write the parmameters for current plot                        
            #--------------------------------------------------------------------------------------
            result_file.write('#' + 60 * '-' + '\n')
            result_file.write('# parameters used in fit \n')            
            result_file.write('#' + 60 * '-' + '\n')
            parms = fit_result.params         
            for p in parms:
                p_name  = p.rsplit('_',1)[0]
                p_n     = int(p.rsplit('_',1)[1])
                if p_n != n_dataset :
                    continue
                result_file.write('# ' + '{:17}'.format(p_name +'_label') + ' : ')
                result_file.write('{:>10.3g}'.format(parms[p].value))
             
                if parms[p_name + '_1'].glbl and i > 0:
                    result_file.write(', (global)\n')
                if not parms[p].vary:
                    result_file.write(', (fixed)\n')
                else:
                    stderr = '{:6.3f}'.format(parms[p].stderr)
                    result_file.write(', (' + unicodedata.lookup('plus-minus sign')
                                       + str(stderr) + ')\n')             

            #--------------------------------------------------------------------------------------# --------------------------------------------------------------------------
            # Get the point where time > 2E-6 sec and ignore earlier points to avoid noise
            #--------------------------------------------------------------------------------------# --------------------------------------------------------------------------
            # i_start = np.where(PlotDataSets[i][0] > 2.0E-6)[0][0]
            #--------------------------------------------------------------------------------------# --------------------------------------------------------------------------
            # Decided it is not a good idea to not print all the data
            # Better to write the plot routine to ignore the data
            #--------------------------------------------------------------------------------------# --------------------------------------------------------------------------
            
            i_start = 0  # maybe not a good idea to throw away data

            #--------------------------------------------------------------------------------------# --------------------------------------------------------------------------
            # Get the plot data and convert time to microseconds
            #--------------------------------------------------------------------------------------# --------------------------------------------------------------------------
            Time = PlotDataSets[i][0][i_start:]
            Signal = PlotDataSets[i][1][i_start:]
            Fit = compute_tof.TOF(Time, n_dataset, fit_result.params, data, glbl)

            mass_factor = np.sqrt(data.mass_molecules[i] / glbl.massH2)
            ion_tof = fit_result.params['ion_tof_%i' % n_dataset].value * mass_factor

            #--------------------------------------------------------------------------------------
            # Compute cutoff for plotting.
            #   Create a new time array which is corrected for ion time of flight
            #   Replace time=0 by time = 1e-8 to remove singularity in velocity and
            #   energy if t=0
            #--------------------------------------------------------------------------------------
            Time2 = Time - ion_tof * 1E-6
            Time2 = np.where(Time2 != 0, Time2, np.repeat(0.01E-6, len(Time)))
            Cutoff = cutoff_function(fit_result.params, data, glbl,
                                            n_dataset, Time2, cutoff_type)
            # Convert time to microseconds for plotting
            Time = Time * 1E6  # convert to microseconds for plotting

            #---------- ---------------------------------------------------------------------------
            # write the number of data lines and range of lines fit
            # -------------------------------------------------------------------------------------
            result_file.write('# n_points          : ' + str(len(Time)) + '\n')
            result_file.write('#' + 68 * '-' + '\n')

            # -------------------------------------------------------------------------------------
            # write the plot data
            # -------------------------------------------------------------------------------------
            result_file.write('# time            Signal                Fit                 Cutoff\n')
            result_file.write('#' + 68 * '-' + '\n')
            result_file.write('# Begin data\n')

            for j in range(len(Time)):
                result_file.write('{:6.2f} {:20.5e} {:20.5e} {:20.5e}\n'
                                  .format(Time[j], Signal[j], Fit[j], Cutoff[j]))

            result_file.write('# End data\n')
            result_file.write('#' + 68 * '-' + '\n')
            result_file.write('# End Plot ' + str(n_dataset) + '\n')

    # end comment out for profiling

    # print("--- %s seconds ---" % (time.time() - start_time))

    return result_filename


# ------------------------------------------------------------------------------
# Enter fit results in results.xlsx
# ------------------------------------------------------------------------------
def write_results_excel_file(glbl, data, path, fit_number, fit_result):
    xls_filename = path + 'fit_results.xlsx'
    # --------------------------------------------------------------------------
    # check if the excel file is open and prompt user to close if needed
    # --------------------------------------------------------------------------
    excel_file_might_be_open = True    
    while excel_file_might_be_open:
        try:
            test_file = open(xls_filename, "r+")
            excel_file_might_be_open = False
            test_file.close
        except IOError:
            input("Please close Excel and hit enter to continue")

    wb = pyxl.load_workbook(xls_filename)
    ws = wb.active

    i_found = None
    next_row = None
    fit_name = 'Fit' + fit_number

    for i in range(1, ws.max_row + 1):
        try:
            if ws.cell(row=i, column=1).value.startswith(fit_name):
                i_found = i
                break
        except:
            pass

    if i_found:
        print('An entry for fit', fit_name, 'already exists')

        while True:
            ans = input('Overwrite (O)  Append (A)  or Skip (S): ').upper()
            if ans.startswith('O'):
                next_row = i_found
                break
            elif ans.startswith('A'):
                next_row = ws.max_row + 1
                break
            elif ans.startswith('S'):
                next_row = None
                break
            else:
                print('Please enter "O", "A", or "S" : ')

    else:
        next_row = ws.max_row + 2
        
    if (next_row):
        for n in range(len(glbl.signal_filenames)):
            
            #--------------------------------------------------------------------------------------
            #  get the actual values of t_min and t_max used in the fits
            #--------------------------------------------------------------------------------------            
            n_min = data.fit_index_ranges[n][0]
            n_max = data.fit_index_ranges[n][1]
            t_min = data.datasets[n][0][n_min] * 1E6
            t_max = data.datasets[n][0][n_max] * 1E6
            
            ws.cell(row=next_row, column=1).value = fit_name + '_' + str(n + 1) # fit name
            if n==0:
                ws.cell(row=next_row, column=2).value = glbl.comment_xlsx       # comment
            ws.cell(row=next_row, column=3).value = data.molecules[n]           # molecule
            ws.cell(row=next_row, column=4).value = data.states[n][0]           # vib state
            ws.cell(row=next_row, column=5).value = data.states[n][1]           # rot state
            ws.cell(row=next_row, column=6).value = fit_result.params['temp_' + str(n + 1)].value
            ws.cell(row=next_row, column=7).value = glbl.functions[n]           # fit function
            ws.cell(row=next_row, column=8).value = glbl.averaging_types[n]     # averaging type
            #ws.cell(row=next_row, column=9).value = '{:8.2e}'.format(fit_result.chisqr) # chi square
            ws.cell(row=next_row, column=9).value = fit_result.chisqr           # chi square
            if glbl.functions[0].lower() == 'erf':
                ws.cell(row=next_row, column=10).value = fit_result.params['e0_' + str(n + 1)].value
                ws.cell(row=next_row, column=11).value = fit_result.params['e0_' + str(n + 1)].stderr
                ws.cell(row=next_row, column=12).value = fit_result.params['w_'  + str(n + 1)].value
                ws.cell(row=next_row, column=13).value = fit_result.params['w_'  + str(n + 1)].stderr
            ws.cell(row=next_row, column=16).value = t_min
            ws.cell(row=next_row, column=17).value = t_max 
            if glbl.fit_ranges[n] == None:
                ws.cell(row=next_row, column=18).value = glbl.thresholds[n]
            ws.cell(row=next_row, column=19).value = fit_result.params['baseline_'  + str(n + 1)].value
            ws.cell(row=next_row, column=20).value = fit_result.params['y_scale_'  + str(n + 1)].value
            ws.cell(row=next_row, column=21).value = fit_result.params['ffr_' + str(n + 1)].value
            ws.cell(row=next_row, column=22).value = fit_result.params['ion_tof_'  + str(n + 1)].value
            if glbl.cutoff_types[n] =='exp':
                ws.cell(row=next_row, column=23).value = fit_result.params['ecutm_' + str(n + 1)].value
                ws.cell(row=next_row, column=24).value = fit_result.params['ecuts_' + str(n + 1)].value
            if glbl.cutoff_types[n] =='tanh':
                ws.cell(row=next_row, column=23).value = fit_result.params['tcutc_' + str(n + 1)].value
                ws.cell(row=next_row, column=24).value = fit_result.params['tcutw_' + str(n + 1)].value
            ws.cell(row=next_row, column=25).value = data.dates[n]
            next_row += 1
    wb.save(xls_filename)
